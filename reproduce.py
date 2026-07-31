#!/usr/bin/env python3
"""Single reproducibility entry point for all programming-stage artifacts."""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from water_quality_model import PROJECT_ROOT, SEED, semantic_workbook_hash, sha256_file


RESULTS = PROJECT_ROOT / "results"
FIGURES = PROJECT_ROOT / "figures"
SKILL_ROOT = PROJECT_ROOT / ".agents" / "skills" / "math-modeling"
REPRODUCE_COMMAND = "PYTHONPATH=.vendor python reproduce.py"
EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")
# Windows 上目录没有 POSIX 写位，只读属性也只对文件生效；输入保护因此只审计文件。
WINDOWS = os.name == "nt"


def relative_path(path: Path) -> str:
    """统一用正斜杠记录仓库内相对路径，使清单跨平台一致。"""
    return path.relative_to(PROJECT_ROOT).as_posix()


def command_environment() -> dict[str, str]:
    environment = os.environ.copy()
    environment["PYTHONPATH"] = ".vendor"
    scratch = Path(tempfile.gettempdir()) / "math-modeling-build"
    mpl_config = scratch / "mplconfig"
    xdg_cache = scratch / "xdg-cache"
    xdg_config = scratch / "xdg-config"
    for directory in (mpl_config, xdg_cache, xdg_config):
        directory.mkdir(parents=True, exist_ok=True)
    environment["MPLCONFIGDIR"] = str(mpl_config)
    environment["XDG_CACHE_HOME"] = str(xdg_cache)
    environment["XDG_CONFIG_HOME"] = str(xdg_config)
    return environment


def run(command: list[str], *, stdout_path: Path | None = None) -> None:
    environment = command_environment()
    if stdout_path is None:
        completed = subprocess.run(command, cwd=PROJECT_ROOT, env=environment, check=False)
    else:
        with stdout_path.open("w", encoding="utf-8") as stream:
            completed = subprocess.run(command, cwd=PROJECT_ROOT, env=environment, stdout=stream, check=False)
    if completed.returncode != 0:
        raise RuntimeError(f"Command failed with exit code {completed.returncode}: {' '.join(command)}")


def validate_input_protection() -> dict:
    data_root = PROJECT_ROOT / "data"
    inputs = sorted(path for path in data_root.rglob("*") if path.is_file())
    directories = [data_root] + sorted(path for path in data_root.rglob("*") if path.is_dir())
    audited = inputs if WINDOWS else [*inputs, *directories]
    records = []
    for path in audited:
        mode = path.stat().st_mode & 0o777
        records.append(
            {
                "path": relative_path(path),
                "kind": "directory" if path.is_dir() else "file",
                "mode": f"{mode:04o}",
                "writable_bits": f"{mode & 0o222:04o}",
            }
        )
    writable = [record for record in records if record["writable_bits"] != "0000"]
    result = {
        "ok": not writable,
        "checked": len(records),
        "directories_audited": not WINDOWS,
        "writable": writable,
        "records": records,
    }
    (RESULTS / "input_protection.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if writable:
        raise AssertionError(f"Input protection failed: {writable}")
    return result


def recalculate_workbook(path: Path) -> dict:
    script = SKILL_ROOT / "tools" / "xlsx" / "scripts" / "recalc.py"
    command = [sys.executable, str(script), str(path), "60"]
    completed = subprocess.run(
        command,
        cwd=PROJECT_ROOT,
        env=command_environment(),
        capture_output=True,
        text=True,
        check=False,
    )
    try:
        result = json.loads(completed.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            f"LibreOffice gate returned invalid JSON (exit {completed.returncode}): "
            f"{completed.stdout}\n{completed.stderr}"
        ) from exc
    if completed.returncode != 0 or result.get("status") != "success":
        raise RuntimeError(
            f"LibreOffice recalculation failed (exit {completed.returncode}): {result}; {completed.stderr}"
        )
    result.update(
        {
            "path": relative_path(path),
            "command": " ".join(command),
            "unique_reproduction_command": REPRODUCE_COMMAND,
            "workbook_sha256": sha256_file(path),
            "workbook_semantic_sha256": semantic_workbook_hash(path),
        }
    )
    (RESULTS / "xlsx_recalc_external.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return result


def validate_workbook(path: Path) -> dict:
    required = (
        "Q1预测",
        "Q1模型评价",
        "Q2时滞与参数",
        "Q2拟合精度",
        "Q3预测",
        "Q3分时距精度",
        "Q3敏感性",
        "Q4等级占比",
        "Q4三月逐日",
        "Q4全期逐日",
        "数据质量",
    )
    workbook = load_workbook(path, read_only=True, data_only=False)
    formula_count = 0
    errors = []
    dimensions = {}
    try:
        for sheet in workbook.worksheets:
            dimensions[sheet.title] = {"rows": sheet.max_row, "columns": sheet.max_column}
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                    if isinstance(value, str) and any(error in value for error in EXCEL_ERRORS):
                        errors.append(f"{sheet.title}!{cell.coordinate}:{value}")
    finally:
        workbook.close()
    missing = [sheet for sheet in required if sheet not in dimensions]
    empty = [sheet for sheet, size in dimensions.items() if size["rows"] < 2]
    result = {
        "ok": not missing and not empty and not errors,
        "path": relative_path(path),
        "sheet_count": len(dimensions),
        "required_missing": missing,
        "empty_sheets": empty,
        "formula_count": formula_count,
        "excel_errors": errors,
        "dimensions": dimensions,
        "libreoffice_rendering": "see results/xlsx_recalc_external.json for the separate sandbox-external gate",
    }
    (RESULTS / "xlsx_validation.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if not result["ok"]:
        raise AssertionError(f"Workbook validation failed: {result}")
    return result


def build_manifest() -> None:
    summary = json.loads((RESULTS / "full_summary.json").read_text(encoding="utf-8"))
    workbook_inputs = sorted((PROJECT_ROOT / "data").rglob("*.xlsx")) + sorted(
        (PROJECT_ROOT / "data").rglob("*.xls")
    )
    all_inputs = [PROJECT_ROOT / "data" / "A题 自来水厂水质预测与评估.pdf", *workbook_inputs]
    semantic_hashes = {
        item["path"]: item["semantic_sha256"] for item in summary["data"]["input_hashes"]
    }
    parameters = {
        "semantic_workbook_sha256": semantic_hashes,
        "q1_ridge": summary["q1"]["best_ridge"],
        "q1_rff": summary["q1"]["best_rff"],
        "q2_delays_steps": summary["q2"]["delays_steps"],
        "q2_ridge": summary["q2"]["ridge"],
        "q2_ar_order": summary["q2"]["selected_ar_order"],
        "q3_tau_hours": summary["q3"]["tau_hours"],
        "q3_ridge_by_horizon": summary["q3"]["ridge_by_horizon"],
        "uncertainty_paths": 500,
        "moving_block_origins": 12,
        "risk_limit_ntu": 1.0,
    }
    command = [
        sys.executable,
        str(SKILL_ROOT / "references" / "roles" / "编程手" / "scripts" / "repro_manifest.py"),
        "--project-root",
        str(PROJECT_ROOT),
    ]
    for path in all_inputs:
        command.extend(("--input", str(path)))
    command.extend(
        (
            "--seed",
            str(SEED),
            "--parameters",
            json.dumps(parameters, ensure_ascii=False, separators=(",", ":")),
            "--command",
            REPRODUCE_COMMAND,
            "--package",
            "numpy",
            "--package",
            "pandas",
            "--package",
            "scipy",
            "--package",
            "matplotlib",
            "--package",
            "openpyxl",
            "--package",
            "xlrd",
            "--output",
            "results/复现清单.json",
            "--overwrite",
        )
    )
    run(command)

    manifest_path = RESULTS / "复现清单.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    code_files = (
        PROJECT_ROOT / "water_quality_model.py",
        PROJECT_ROOT / "water_quality_full.py",
        PROJECT_ROOT / "water_quality_figures.py",
        PROJECT_ROOT / "reproduce.py",
        PROJECT_ROOT / "utils" / "plot_style.py",
    )
    output_files = sorted(
        path
        for root in (RESULTS, FIGURES)
        for path in root.rglob("*")
        if path.is_file() and path != manifest_path and "_qa" not in path.parts
    )
    manifest["semantic_workbook_sha256"] = semantic_hashes
    # 数据源唯一化的等价性判定：仓库曾并存两份 A 题附件，副本经语义哈希逐文件核对后删除。
    # 副本删除后无法再重算，故把判定结果长期绑定在清单里。
    equivalence_path = RESULTS / "数据源等价性判定.json"
    if equivalence_path.exists():
        equivalence = json.loads(equivalence_path.read_text(encoding="utf-8"))
        manifest["data_source_deduplication"] = {
            key: equivalence[key]
            for key in ("判定时间", "结论", "判据", "文件对数", "语义一致数", "字节一致数", "说明")
        }
        manifest["data_source_deduplication"]["明细"] = relative_path(equivalence_path)
    manifest["code_files"] = [
        {"path": relative_path(path), "sha256": sha256_file(path), "bytes": path.stat().st_size}
        for path in code_files
    ]
    manifest["output_files"] = [
        {"path": relative_path(path), "sha256": sha256_file(path), "bytes": path.stat().st_size}
        for path in output_files
    ]
    manifest["output_workbook_semantic_sha256"] = semantic_workbook_hash(RESULTS / "建模结果.xlsx")
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    RESULTS.mkdir(parents=True, exist_ok=True)
    validate_input_protection()
    run([sys.executable, "water_quality_full.py", "--full"])
    run([sys.executable, "water_quality_figures.py", "--all"])
    run(
        [
            sys.executable,
            str(SKILL_ROOT / "references" / "roles" / "编程手" / "scripts" / "figure_audit.py"),
            "figures",
            "--questions",
            "q1",
            "q2",
            "q3",
            "q4",
            "--strict",
        ],
        stdout_path=RESULTS / "figure_audit.json",
    )
    validate_workbook(RESULTS / "建模结果.xlsx")
    recalculate_workbook(RESULTS / "建模结果.xlsx")
    validate_workbook(RESULTS / "建模结果.xlsx")
    build_manifest()
    print(json.dumps({"status": "ok", "command": REPRODUCE_COMMAND}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
