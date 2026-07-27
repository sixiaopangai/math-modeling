import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


SCRIPTS = Path(__file__).resolve().parents[1] / "tools" / "xlsx" / "scripts"
sys.path.insert(0, str(SCRIPTS))

from read_rows import read_excel_rows


class ExcelReadTests(unittest.TestCase):
    def test_headerless_sheet_keeps_first_value_and_all_7470_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "spectrum.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.cell(1, 1, 399.6747)
            for row in range(2, 7471):
                sheet.cell(row, 1, float(row))
            workbook.save(path)

            rows = read_excel_rows(path, header=False, expected_rows=7470)

        self.assertEqual(len(rows), 7470)
        self.assertEqual(rows[0][0], 399.6747)

    def test_expected_row_count_mismatch_fails_loudly(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "data.xlsx"
            workbook = Workbook()
            workbook.active.append([1])
            workbook.save(path)

            with self.assertRaisesRegex(ValueError, "期望 2 行"):
                read_excel_rows(path, header=False, expected_rows=2)


if __name__ == "__main__":
    unittest.main()
