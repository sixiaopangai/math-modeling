#!/usr/bin/env python3
"""Water-quality modeling pipeline for the 2026 APMCM problem A."""

from __future__ import annotations

import argparse
import hashlib
import itertools
import json
import math
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent
VENDOR = PROJECT_ROOT / ".vendor"
if VENDOR.is_dir():
    sys.path.insert(0, str(VENDOR))

import numpy as np
import pandas as pd
import xlrd
from openpyxl import load_workbook

SEED = 20260727
STEP_HOURS = 2
MODEL_LAMBDAS = (0.01, 0.1, 1.0, 10.0, 100.0)
TIME_SLOTS = (700, 900, 1100, 1300, 1500, 1700, 1900, 2100, 2300, 100, 300, 500)

FIELDS = (
    "TIME",
    "RIVER LEVEL",
    "R/W PUMP DUTY",
    "R/W FLOW",
    "R/W NTU",
    "R/W CLR",
    "R/W PH",
    "FILT. NTU",
    "C/W WELL LEVEL",
    "PH",
    "NTU",
    "CLR",
    "CL2",
    "F/RIDE",
    "ALUM",
    "T/W PUMP DUTY",
    "T/W FLOW",
    "18ML LEVEL",
    "18ML FLOW",
    "REMARKS",
)

NUMERIC_NAMES = {
    "RIVER LEVEL": "river_level",
    "R/W PUMP DUTY": "rw_pump_duty",
    "R/W FLOW": "rw_flow",
    "R/W NTU": "rw_ntu",
    "R/W CLR": "rw_clr",
    "R/W PH": "rw_ph",
    "FILT. NTU": "filt_ntu",
    "C/W WELL LEVEL": "cw_level",
    "PH": "treated_ph",
    "NTU": "ntu",
    "CLR": "treated_clr",
    "CL2": "cl2",
    "F/RIDE": "f_ride",
    "ALUM": "alum",
    "T/W FLOW": "tw_flow",
    "18ML LEVEL": "tank18_level",
    "18ML FLOW": "tank18_flow",
}

MONTHS = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "July": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().upper()
    aliases = {"DATA": "DATE", "RIVER LEVEL": "RIVER LEVEL"}
    return aliases.get(text, text)


def parse_numeric(value: Any) -> tuple[float, bool]:
    if value is None or (isinstance(value, str) and value.strip() in ("", "-")):
        return np.nan, False
    try:
        return float(str(value).strip()), False
    except (TypeError, ValueError):
        return np.nan, True


def normalize_time(value: Any) -> str:
    try:
        return f"{int(float(value)):04d}"
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid time value: {value!r}") from exc


def make_timestamp(op_day: date, time_value: Any) -> datetime:
    text = normalize_time(time_value)
    hour, minute = int(text[:2]), int(text[2:])
    calendar_day = op_day + (timedelta(days=1) if hour < 7 else timedelta())
    return datetime.combine(calendar_day, datetime.min.time()).replace(hour=hour, minute=minute)


def semantic_workbook_hash(path: Path) -> str:
    digest = hashlib.sha256()

    def update(sheet: str, values: list[Any]) -> None:
        normalized = []
        for value in values:
            if value is None:
                normalized.append(None)
            elif isinstance(value, float):
                normalized.append(format(value, ".15g"))
            elif hasattr(value, "isoformat"):
                normalized.append(value.isoformat())
            else:
                normalized.append(str(value).strip())
        payload = json.dumps([sheet, normalized], ensure_ascii=False, separators=(",", ":"))
        digest.update(payload.encode("utf-8"))

    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                update(sheet.title, list(row))
    else:
        workbook = xlrd.open_workbook(str(path))
        for sheet in workbook.sheets():
            for row_index in range(sheet.nrows):
                update(sheet.name, sheet.row_values(row_index))
    return digest.hexdigest()


def row_record(
    row: list[Any] | tuple[Any, ...],
    mapping: dict[str, int],
    op_day: date,
    source: str,
    expected_time: int,
) -> dict[str, Any]:
    time_value = row[mapping["TIME"]]
    observed_time = normalize_time(time_value)
    normalized_expected = normalize_time(expected_time)
    record: dict[str, Any] = {
        "timestamp": make_timestamp(op_day, expected_time),
        "op_day": op_day,
        "source": source,
        "time": normalized_expected,
        "time_raw": str(time_value).strip(),
        "time_repaired": observed_time != normalized_expected,
    }
    for field, output_name in NUMERIC_NAMES.items():
        raw = row[mapping[field]] if field in mapping else None
        numeric, invalid = parse_numeric(raw)
        record[output_name] = numeric
        record[f"{output_name}_raw"] = raw
        record[f"{output_name}_clean"] = numeric
        record[f"{output_name}_invalid"] = bool(invalid)
    raw_pump_duty = row[mapping["R/W PUMP DUTY"]] if "R/W PUMP DUTY" in mapping else None
    record["rw_pump_duty_category"] = None if raw_pump_duty is None else str(raw_pump_duty).strip()
    for field, output_name in (("T/W PUMP DUTY", "tw_pump_duty"), ("REMARKS", "remarks")):
        record[output_name] = row[mapping[field]] if field in mapping else None
    return record


def load_data(project_root: Path = PROJECT_ROOT) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    data_root = project_root / "data"
    xlsx_files = sorted(data_root.rglob("*.xlsx"))
    xls_files = sorted(data_root.rglob("*.xls"))
    if len(xlsx_files) != 12 or len(xls_files) != 3:
        raise AssertionError(f"Expected 12 xlsx and 3 xls files, got {len(xlsx_files)} and {len(xls_files)}")

    records: list[dict[str, Any]] = []
    inputs: list[dict[str, Any]] = []
    allowed = set(FIELDS) | {"DATE"}

    for path in xlsx_files:
        month = next((number for label, number in MONTHS.items() if label in path.stem), None)
        if month is None:
            raise ValueError(f"Cannot infer month from {path.name}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator)
        mapping = {
            normalized: index
            for index, value in enumerate(header)
            if (normalized := normalize_header(value)) is not None and normalized != "DATE"
        }
        unknown = set(mapping) - allowed
        if unknown or "TIME" not in mapping:
            raise AssertionError(f"Unexpected schema in {path.name}: {sorted(unknown)}")
        rows = [list(row) for row in iterator if row[mapping["TIME"]] not in (None, "")]
        expected = pd.Period(f"2025-{month:02d}").days_in_month * 12
        if len(rows) != expected:
            raise AssertionError(f"{path.name}: expected {expected} records, got {len(rows)}")
        for index, row in enumerate(rows):
            op_day = date(2025, month, 1) + timedelta(days=index // 12)
            records.append(row_record(row, mapping, op_day, path.name, TIME_SLOTS[index % 12]))
        inputs.append(
            {
                "path": path.relative_to(project_root).as_posix(),
                "sha256": sha256_file(path),
                "semantic_sha256": semantic_workbook_hash(path),
                "records": len(rows),
            }
        )

    for path in xls_files:
        workbook = xlrd.open_workbook(str(path))
        file_records = 0
        for sheet in workbook.sheets():
            match = re.match(r"\s*(\d{1,2})\.(\d{1,2})", sheet.name)
            if not match:
                raise AssertionError(f"Unexpected sheet name {sheet.name!r} in {path.name}")
            op_day = date(2026, int(match.group(2)), int(match.group(1)))
            header = sheet.row_values(0)
            mapping = {
                normalized: index
                for index, value in enumerate(header)
                if (normalized := normalize_header(value)) is not None
            }
            unknown = set(mapping) - allowed
            if unknown or set(FIELDS) - set(mapping):
                raise AssertionError(f"Unexpected schema in {path.name}/{sheet.name}: {sorted(unknown)}")
            if sheet.nrows != 13:
                raise AssertionError(f"{path.name}/{sheet.name}: expected 13 physical rows, got {sheet.nrows}")
            for row_index in range(1, sheet.nrows):
                records.append(
                    row_record(sheet.row_values(row_index), mapping, op_day, path.name, TIME_SLOTS[row_index - 1])
                )
                file_records += 1
        expected_sheets = {1: 31, 2: 28, 3: 31}[int(re.search(r"(\d+)月", path.stem).group(1))]
        if workbook.nsheets != expected_sheets or file_records != expected_sheets * 12:
            raise AssertionError(f"{path.name}: unexpected sheets/records")
        inputs.append(
            {
                "path": path.relative_to(project_root).as_posix(),
                "sha256": sha256_file(path),
                "semantic_sha256": semantic_workbook_hash(path),
                "records": file_records,
            }
        )

    frame = pd.DataFrame.from_records(records).sort_values("timestamp").reset_index(drop=True)
    if len(frame) != 5460 or frame["timestamp"].nunique() != 5460:
        raise AssertionError("Timeline must contain 5460 unique timestamps")
    deltas = frame["timestamp"].diff().dropna().dt.total_seconds().div(3600)
    if not np.allclose(deltas.to_numpy(), STEP_HOURS):
        bad = frame.loc[deltas.index[~np.isclose(deltas.to_numpy(), STEP_HOURS)], ["timestamp", "source"]]
        raise AssertionError(
            "Timeline is not uniformly spaced at two hours: "
            f"delta_counts={deltas.value_counts().sort_index().to_dict()}, "
            f"first_bad={bad.head(5).to_dict(orient='records')}"
        )
    return frame, inputs


@dataclass
class RidgeModel:
    coef: np.ndarray
    median: np.ndarray
    lower: np.ndarray
    upper: np.ndarray
    mean: np.ndarray
    scale: np.ndarray

    def transform(self, values: np.ndarray) -> np.ndarray:
        data = np.asarray(values, dtype=float).copy()
        data = np.where(np.isnan(data), self.median, data)
        data = np.clip(data, self.lower, self.upper)
        return (data - self.mean) / self.scale

    def predict(self, values: np.ndarray) -> np.ndarray:
        data = self.transform(values)
        return self.coef[0] + data @ self.coef[1:]


def fit_ridge(values: np.ndarray, target: np.ndarray, ridge: float, robust: bool = False) -> RidgeModel:
    values = np.asarray(values, dtype=float)
    target = np.asarray(target, dtype=float)
    median = np.nanmedian(values, axis=0)
    lower = np.nanpercentile(values, 0.5, axis=0)
    upper = np.nanpercentile(values, 99.5, axis=0)
    clean = np.where(np.isnan(values), median, values)
    clean = np.clip(clean, lower, upper)
    mean = np.mean(clean, axis=0)
    scale = np.std(clean, axis=0)
    scale = np.where(scale < 1e-10, 1.0, scale)
    standardized = (clean - mean) / scale
    design = np.column_stack((np.ones(len(standardized)), standardized))
    penalty = np.eye(design.shape[1]) * ridge
    penalty[0, 0] = 0.0
    weights = np.ones(len(target))
    coef = np.zeros(design.shape[1])
    iterations = 5 if robust else 1
    for _ in range(iterations):
        weighted = design * np.sqrt(weights)[:, None]
        response = target * np.sqrt(weights)
        coef = np.linalg.solve(weighted.T @ weighted + penalty, weighted.T @ response)
        if robust:
            residual = target - design @ coef
            center = np.median(residual)
            mad = np.median(np.abs(residual - center))
            robust_scale = max(1.4826 * mad, 1e-8)
            score = np.abs(residual / robust_scale)
            weights = np.where(score <= 1.345, 1.0, 1.345 / np.maximum(score, 1e-12))
    return RidgeModel(coef, median, lower, upper, mean, scale)


def metric_dict(actual: np.ndarray, predicted: np.ndarray) -> dict[str, float]:
    actual = np.asarray(actual, dtype=float)
    predicted = np.asarray(predicted, dtype=float)
    rmse = float(np.sqrt(np.mean((actual - predicted) ** 2)))
    mae = float(np.mean(np.abs(actual - predicted)))
    denominator = float(np.sum((actual - np.mean(actual)) ** 2))
    r2 = float(1.0 - np.sum((actual - predicted) ** 2) / denominator) if denominator > 0 else float("nan")
    return {"rmse": rmse, "mae": mae, "r2": r2}


def q1_feature_frame(frame: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    variables = [
        "river_level",
        "rw_flow",
        "rw_ntu",
        "rw_clr",
        "rw_ph",
        "filt_ntu",
        "cw_level",
        "treated_ph",
        "cl2",
        "alum",
        "tw_flow",
    ]
    features: dict[str, pd.Series] = {}
    for variable in variables:
        base = np.log1p(frame[variable]) if variable in {"rw_ntu", "rw_clr", "filt_ntu"} else frame[variable]
        features[variable] = base
        for lag in (1, 2, 3, 6):
            features[f"{variable}_lag{lag}"] = base.shift(lag)
    hour = frame["timestamp"].dt.hour + frame["timestamp"].dt.minute / 60
    day = frame["timestamp"].dt.dayofyear
    features["hour_sin"] = np.sin(2 * np.pi * hour / 24)
    features["hour_cos"] = np.cos(2 * np.pi * hour / 24)
    features["year_sin"] = np.sin(2 * np.pi * day / 365.25)
    features["year_cos"] = np.cos(2 * np.pi * day / 365.25)
    result = pd.DataFrame(features)
    return result, list(result.columns)


def fit_q1_smoke(frame: pd.DataFrame) -> dict[str, Any]:
    features, names = q1_feature_frame(frame)
    target = np.log1p(frame["ntu"])
    train = (frame["timestamp"] < "2026-01-01") & target.notna()
    validation = (frame["timestamp"] >= "2026-01-01") & (frame["timestamp"] < "2026-02-01") & target.notna()
    scores = []
    for ridge in MODEL_LAMBDAS:
        model = fit_ridge(features.loc[train].to_numpy(), target.loc[train].to_numpy(), ridge, robust=True)
        prediction = np.maximum(0.0, np.expm1(model.predict(features.loc[validation].to_numpy())))
        scores.append((metric_dict(frame.loc[validation, "ntu"].to_numpy(), prediction)["rmse"], ridge))
    best_ridge = min(scores)[1]
    model = fit_ridge(features.loc[train].to_numpy(), target.loc[train].to_numpy(), best_ridge, robust=True)
    prediction = np.maximum(0.0, np.expm1(model.predict(features.loc[validation].to_numpy())))
    metrics = metric_dict(frame.loc[validation, "ntu"].to_numpy(), prediction)
    return {
        "train_rows": int(train.sum()),
        "validation_rows": int(validation.sum()),
        "feature_count": len(names),
        "ridge": best_ridge,
        "metrics_ntu": metrics,
        "prediction_min": float(np.min(prediction)),
        "prediction_max": float(np.max(prediction)),
    }


def q2_design(frame: pd.DataFrame, delays: tuple[int, int, int, int]) -> tuple[pd.DataFrame, pd.Series]:
    target = np.log1p(frame["filt_ntu"])
    data: dict[str, pd.Series] = {f"target_lag{lag}": target.shift(lag) for lag in (1, 2, 3)}
    variables = ("rw_ntu", "rw_ph", "alum", "rw_flow")
    for variable, delay in zip(variables, delays):
        series = frame[variable].shift(delay)
        if variable == "rw_ntu":
            series = np.log1p(series)
        data[variable] = series
        data[f"{variable}_sq"] = series**2
    data["ntu_alum"] = np.log1p(frame["rw_ntu"].shift(delays[0])) * frame["alum"].shift(delays[2])
    data["ph_alum"] = frame["rw_ph"].shift(delays[1]) * frame["alum"].shift(delays[2])
    return pd.DataFrame(data), target


def fit_q2_smoke(frame: pd.DataFrame) -> dict[str, Any]:
    best: tuple[float, tuple[int, int, int, int], float, dict[str, float]] | None = None
    for delays in itertools.product(range(4), repeat=4):
        design, target = q2_design(frame, delays)
        complete = design.notna().all(axis=1) & target.notna()
        train = complete & (frame["timestamp"] < "2026-01-01")
        validation = complete & (frame["timestamp"] >= "2026-01-01") & (frame["timestamp"] < "2026-02-01")
        if train.sum() < 100 or validation.sum() < 50:
            continue
        for ridge in MODEL_LAMBDAS:
            model = fit_ridge(design.loc[train].to_numpy(), target.loc[train].to_numpy(), ridge)
            prediction = np.maximum(0.0, np.expm1(model.predict(design.loc[validation].to_numpy())))
            metrics = metric_dict(frame.loc[validation, "filt_ntu"].to_numpy(), prediction)
            key = (metrics["rmse"], sum(delays), delays, ridge)
            if best is None or key < (best[0], sum(best[1]), best[1], best[2]):
                best = (metrics["rmse"], delays, ridge, metrics)
    if best is None:
        raise RuntimeError("Q2 lag search found no valid model")
    return {
        "delay_steps": dict(zip(("rw_ntu", "rw_ph", "alum", "rw_flow"), best[1])),
        "delay_hours": dict(zip(("rw_ntu", "rw_ph", "alum", "rw_flow"), (2 * x for x in best[1]))),
        "ridge": best[2],
        "validation_metrics_ntu": best[3],
    }


def rtd_kernel(tau_hours: float) -> np.ndarray:
    q = math.exp(-STEP_HOURS / tau_hours)
    length = math.ceil(math.log(0.05) / math.log(q))
    indices = np.arange(length)
    weights = q**indices
    weights /= np.sum(weights)
    if 1 - q**length < 0.95 - 1e-12 or not np.isclose(np.sum(weights), 1.0):
        raise AssertionError("RTD mass constraint failed")
    return weights


def convolve_rtd(series: pd.Series, tau_hours: float) -> np.ndarray:
    values = series.interpolate(limit=3).ffill().bfill().to_numpy(dtype=float)
    return np.convolve(values, rtd_kernel(tau_hours), mode="full")[: len(values)]


def one_step_physical(frame: pd.DataFrame, origin: int, kernel: np.ndarray) -> float:
    filtered = frame["filt_ntu"].to_numpy(dtype=float)
    last = filtered[origin]
    if np.isnan(last):
        history = filtered[: origin + 1]
        last = history[~np.isnan(history)][-1]
    total = kernel[0] * last
    for lag in range(1, len(kernel)):
        index = origin - (lag - 1)
        value = filtered[index] if index >= 0 else filtered[0]
        if np.isnan(value):
            value = last
        total += kernel[lag] * value
    return float(total)


def fit_q3_smoke(frame: pd.DataFrame) -> dict[str, Any]:
    train_mask = (frame["timestamp"] < "2026-01-01") & frame["ntu"].notna()
    validation_mask = (frame["timestamp"] >= "2026-01-01") & (frame["timestamp"] < "2026-02-01") & frame["ntu"].notna()
    tau_scores = []
    for tau in range(2, 25, 2):
        physical = convolve_rtd(frame["filt_ntu"], tau)
        predictor = np.log1p(physical)[:, None]
        model = fit_ridge(predictor[train_mask], np.log1p(frame.loc[train_mask, "ntu"].to_numpy()), 0.1)
        predicted = np.maximum(0.0, np.expm1(model.predict(predictor[validation_mask])))
        tau_scores.append((metric_dict(frame.loc[validation_mask, "ntu"].to_numpy(), predicted)["rmse"], tau))
    tau = min(tau_scores)[1]
    physical = convolve_rtd(frame["filt_ntu"], tau)
    residual = np.log1p(frame["ntu"]) - np.log1p(physical)
    kernel = rtd_kernel(tau)
    features = []
    targets = []
    target_values = []
    origins = []
    for origin in range(6, len(frame) - 1):
        state = residual.iloc[origin - 5 : origin + 1].to_numpy(dtype=float)[::-1]
        current = frame.loc[origin, ["rw_ntu", "filt_ntu", "alum", "rw_flow", "tw_flow"]].to_numpy(dtype=float)
        hour = frame.loc[origin, "timestamp"].hour
        row = np.concatenate(
            [state, np.array([np.log1p(current[0]), np.log1p(current[1]), *current[2:], math.sin(2 * math.pi * hour / 24), math.cos(2 * math.pi * hour / 24)])]
        )
        target_index = origin + 1
        if np.isnan(residual.iloc[target_index]) or np.isnan(state).any():
            continue
        future_physical = one_step_physical(frame, origin, kernel)
        features.append(row)
        targets.append(math.log1p(frame.loc[target_index, "ntu"]) - math.log1p(future_physical))
        target_values.append(frame.loc[target_index, "ntu"])
        origins.append(origin)
    features_array = np.asarray(features)
    target_array = np.asarray(targets)
    target_values_array = np.asarray(target_values)
    timestamps = frame.loc[np.asarray(origins) + 1, "timestamp"].reset_index(drop=True)
    train = (timestamps < pd.Timestamp("2026-01-01")).to_numpy()
    validation = ((timestamps >= pd.Timestamp("2026-01-01")) & (timestamps < pd.Timestamp("2026-02-01"))).to_numpy()
    best = None
    for ridge in MODEL_LAMBDAS:
        model = fit_ridge(features_array[train], target_array[train], ridge, robust=True)
        residual_prediction = model.predict(features_array[validation])
        prediction = []
        validation_origins = np.asarray(origins)[validation]
        for value, origin in zip(residual_prediction, validation_origins):
            prediction.append(max(0.0, math.expm1(math.log1p(one_step_physical(frame, int(origin), kernel)) + value)))
        metrics = metric_dict(target_values_array[validation], np.asarray(prediction))
        candidate = (metrics["rmse"], ridge, metrics, prediction)
        if best is None or candidate[0] < best[0]:
            best = candidate
    return {
        "tau_hours": tau,
        "kernel_steps": len(kernel),
        "kernel_mass": float(np.sum(kernel)),
        "ridge": best[1],
        "validation_metrics_ntu": best[2],
        "prediction_min": float(np.min(best[3])),
        "prediction_max": float(np.max(best[3])),
    }


def classify_day(values: np.ndarray) -> dict[str, Any]:
    values = np.asarray(values, dtype=float)
    values = values[~np.isnan(values)]
    if len(values) < 9:
        return {"grade": "INSUFFICIENT", "valid_points": int(len(values))}
    excess = np.maximum(values - 1.0, 0.0)
    run = longest = 0
    for flag in excess > 0:
        run = run + 1 if flag else 0
        longest = max(longest, run)
    amplitude = float(np.max(excess))
    duration = float(longest * STEP_HOURS)
    load = float(np.sum(excess) * STEP_HOURS)
    if amplitude == 0:
        grade = "SAFE"
    elif amplitude <= 0.5 and duration <= 4:
        grade = "LOW"
    elif amplitude <= 1.0 and duration <= 8:
        grade = "MEDIUM"
    else:
        grade = "HIGH"
    return {
        "grade": grade,
        "valid_points": int(len(values)),
        "max_ntu": float(np.max(values)),
        "max_excess_ntu": amplitude,
        "longest_exceedance_hours": duration,
        "exceedance_load_ntu_hours": load,
    }


def run_smoke(project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    frame, inputs = load_data(project_root)
    invalid_counts = {
        name: int(frame[f"{name}_invalid"].sum())
        for name in NUMERIC_NAMES.values()
        if int(frame[f"{name}_invalid"].sum()) > 0
    }
    february = (frame["op_day"] >= date(2026, 2, 1)) & (frame["op_day"] <= date(2026, 2, 28))
    time_repairs = frame.loc[frame["time_repaired"], ["source", "op_day", "time_raw", "time"]]
    if len(time_repairs) != 5:
        raise AssertionError(f"Expected 5 auditable time-cell repairs, got {len(time_repairs)}")
    if int(february.sum()) != 336 or int(frame.loc[february, "ntu"].isna().sum()) != 336:
        raise AssertionError("February 2026 must contain 336 rows with all treated-water NTU values missing")
    if invalid_counts.get("rw_flow") != 1 or invalid_counts.get("tw_flow") != 1:
        raise AssertionError("Known invalid R/W FLOW and T/W FLOW cells were not identified")
    q4_sample = frame[frame["op_day"] == date(2026, 1, 10)]["ntu"].to_numpy()
    result = {
        "mode": "smoke",
        "seed": SEED,
        "data": {
            "rows": len(frame),
            "unique_timestamps": int(frame["timestamp"].nunique()),
            "start": frame["timestamp"].min().isoformat(),
            "end": frame["timestamp"].max().isoformat(),
            "february_2026_rows": int(february.sum()),
            "february_2026_ntu_missing": int(frame.loc[february, "ntu"].isna().sum()),
            "invalid_numeric_counts": invalid_counts,
            "time_repairs": [
                {
                    "source": row.source,
                    "op_day": row.op_day.isoformat(),
                    "raw": row.time_raw,
                    "corrected": row.time,
                }
                for row in time_repairs.itertuples(index=False)
            ],
            "inputs": inputs,
        },
        "q1": fit_q1_smoke(frame),
        "q2": fit_q2_smoke(frame),
        "q3": fit_q3_smoke(frame),
        "q4_sample_2026_01_10": classify_day(q4_sample),
    }
    output = project_root / "results" / "p1_smoke.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--smoke", action="store_true", help="Run the P1 vertical slice")
    args = parser.parse_args()
    if not args.smoke:
        parser.error("Use --smoke; full mode is enabled only after P1 passes")
    result = run_smoke()
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
